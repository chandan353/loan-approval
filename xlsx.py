from openpyxl import load_workbook
import mysql.connector
from datetime import datetime

workbook = load_workbook('loan_data.xlsx')
sheet = workbook.active

headers = [str(cell.value) for cell in sheet[1]]

connection = mysql.connector.connect(
    host='localhost',
    user='root',
    password='root',
    database='alemeno',
    auth_plugin='mysql_native_password'
)

cursor = connection.cursor()

try:
    for row in sheet.iter_rows(min_row=2, values_only=True):

        # Add 'created_at' and 'updated_at' to the row
        row += (datetime.now(), datetime.now())

        sql = 'INSERT INTO customer_loan_data ({}) VALUES ({})'.format(
            ', '.join(headers + ['created_at', 'updated_at']),
            ', '.join(['%s'] * len(row))
        )

        cursor.execute(sql, row)
        connection.commit()
        print('Row inserted')

finally:
    cursor.close()
    connection.close()
    print('MySQL connection closed')
